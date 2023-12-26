import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import concurrent.futures
import secrets
import tls_client
import re
import urllib3

import random
# Disable all warnings from urllib3
urllib3.disable_warnings()



  

def trim_and_pad(number):
    # Convert the number to a string and remove the last character
    trimmed_number = str(number)[:-1]
    # Calculate how many zeros to add at the start
    zeros_to_add = 13 - len(trimmed_number)
    # Add the required number of zeros at the start
    modified_number = str('0' * zeros_to_add) + str(trimmed_number)
    return str(modified_number)


ip = {
    'http': 'http://aimleap:b7de10-84d02c-25153e-d60b24-a59354@usa.rotating.proxyrack.net:9000',
    'https': 'http://aimleap:b7de10-84d02c-25153e-d60b24-a59354@usa.rotating.proxyrack.net:9000'
}


# Define the list of items to check
items_to_check = ['crv', 'oz', 'fz', 'count', 'ounce', 'fl',
                  'gallon', 'ct', 'lb', 'liter', 'lt', 'pound', 'kg', 'ml']

items_to_check = [item[::-1] for item in items_to_check]

print(items_to_check)

# Create a regex pattern to match any number followed by an element from the list in reverse
pattern = r'(' + '|'.join(re.escape(item)
                          for item in items_to_check) + r')\s*(\d+(?:\.\d+)?)'
# Function to find all matches in a reversed string


def find_matches_reverse(text):
    reversed_text = text[::-1]  # Reverse the input string
    matches = re.findall(pattern, reversed_text, re.IGNORECASE)
    return matches

cookies = {
    'TealeafAkaSid': 'MgMjEvN2FU2Vkn6RjDYRXv1GLpwRJ6dP',
    'visitorId': '018B4718A59302019EC544D7283F1E05',
    'sapphire': '1',
    '_gcl_au': '1.1.960206792.1697705207',
    'crl8.fpcuid': '74492362-20c1-463a-b7c8-a7b3c8b7ea8f',
    '_ga': 'GA1.2.782197828.1699168442',
    '_ga_N15GRPKXME': 'GS1.1.1699168442.1.1.1699168470.32.0.0',
    '_ce.s': 'v~5201ee79cc2bbcb3e7c02797510daf3bfcaa558e~lcw~1699168470226~lva~1699168442828~vpv~0~v11.fhb~1699168443768~v11.lhb~1699168449039~v11.cs~316943~v11.s~e68d22d0-7baa-11ee-9fe2-fb73ddc9c3f0~v11.sla~1699168470573~v11.send~1699168453916~lcw~1699168470573',
    'fiatsCookie': 'DSI_3314|DSN_Chicago%20Mayfair|DSZ_60630',
    'UserLocation': '60630|41.96576195432252|-87.76362898353945|IL|US',
    # 'ffsession': '{%22sessionHash%22:%2216f5e9a1a7f0941701156617230%22}',
    # 'egsSessionId': '783701f6-32a3-427e-8941-57070595b05c',
    # 'accessToken': 'eyJraWQiOiJlYXMyIiwiYWxnIjoiUlMyNTYifQ.eyJzdWIiOiIxNWM2YzFhMC02OTdlLTQ3MmQtYTdkYS05NWM5ZjA5OWJiMjEiLCJpc3MiOiJNSTYiLCJleHAiOjE3MDEyNDMwMTgsImlhdCI6MTcwMTE1NjYxOCwianRpIjoiVEdULmJhOGE5YjJlMmI1MzQ5NmJhMWU1ZjNlYzY3OWY1YjNmLWwiLCJza3kiOiJlYXMyIiwic3V0IjoiRyIsImRpZCI6IjgxMGE3ZmNmNzVkMzIxZmZjYzExOTk3MDk4N2Y0OTg2Yjc0ZTM2ZTA1ZjI4NjRmOGFiOWQzMDI1YmFjMTgyMGQiLCJzY28iOiJlY29tLm5vbmUsb3BlbmlkIiwiY2xpIjoiZWNvbS13ZWItMS4wLjAiLCJhc2wiOiJMIn0.An58xaAGm0suYZzZrYOt672755XqUsg8QirsrlRvX7NPmnQpC-UAP0pu6g5nRQQAjA4NEgAp7CjSeSG2fZy0HXWNFc6g3_2pmammszNPQ2mEzzPkXRVTH7eNV0vQKMr-o45D--biZFoghOsjY4zvZ6b7mZf0Gv3QU8EUrdFtzKuYXiul7cSBCMOxA9GyQ3M6ct7c2-A6Vi0TOMjCCIUwnW5Tp0i5audM6WCje2T-AWFMqVUTe1zmZX4qmFXo2vFPt7n8sl03r84XcrWZjzJI2cob-emCUZq2eBEszT77k9lVRzRfc9IP_z1Uk83z2LCKHonRkbtEQ01cRWPDzp3j0A',
    # 'idToken': 'eyJhbGciOiJub25lIn0.eyJzdWIiOiIxNWM2YzFhMC02OTdlLTQ3MmQtYTdkYS05NWM5ZjA5OWJiMjEiLCJpc3MiOiJNSTYiLCJleHAiOjE3MDEyNDMwMTgsImlhdCI6MTcwMTE1NjYxOCwiYXNzIjoiTCIsInN1dCI6IkciLCJjbGkiOiJlY29tLXdlYi0xLjAuMCIsInBybyI6eyJmbiI6bnVsbCwiZW0iOm51bGwsInBoIjpmYWxzZSwibGVkIjpudWxsLCJsdHkiOmZhbHNlfX0.',
    # 'refreshToken': '5nl35I6sE-uHh70VlAt_Gd0mfjubQO_VwI_yiYTy-E3-YD21M-Hq0Phwb5g-rgB_-yHKMezkEFKRMarRYZ-JQg',
    '__gads': 'ID=0e70dc00619d5d00:T=1697704892:RT=1701156618:S=ALNI_MbCWts2qdevbDmidCpzranB7PuKXA',
    '__gpi': 'UID=00000c6759d3a73e:T=1697704892:RT=1701156618:S=ALNI_MaiO-BZig0n2q9zCkN0YnkwQ8_cJg',
    'sddStore': 'DSI_3314|DSN_Chicago%20Mayfair|DSZ_60630',
    '_mitata': 'NzlmNjE0Y2MwNzZlNDc3NDk2YmNmMWIxZDFmYjk4YWUzMDBmZGQ4YTk2MjBlYWExZjE3MTU3MDc2NThmZDQ5Mw==_/@#/1701156680_/@#/cBHaMJGxclULbSy6_/@#/NGI2OGE3OGFlZDg0YzIzMzFlYTM1NjNiMGQzNGI3YjFmNTQwMzhlMzI3OWJhYmJlMTYwMjdiNDg1ZjZlY2Y2YQ==_/@#/000',
}

def pro():
    proxy_data = [
        "residential.proxyomega.com:10000:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10001:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10002:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10003:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10004:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac"
    ]

    # Create a list of tuples from the data
    proxies = [tuple(proxy.split(':')) for proxy in proxy_data]

    # Randomly select a proxy from the list
    random_proxy = random.choice(proxies)
    proxy_url = f"http://{random_proxy[2]}:{random_proxy[3]}@{random_proxy[0]}:{random_proxy[1]}"

    # Set up the proxy dictionary
    proxy = {
        'http': proxy_url,
        'https': proxy_url,
    }
    return proxy


def generate_random_hex_string(length):
    # Generate a random sequence of bytes
    random_bytes = secrets.token_bytes(length // 2)
    # Convert the bytes to a hexadecimal string and make it uppercase
    random_hex_string = secrets.token_hex(length // 2).upper()
    return random_hex_string


headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control": "max-age=0",
    "If-None-Match": "\"9cdsdmmez73vfs\"",
    "Sec-Ch-Ua": "\"Chromium\";v=\"116\", \"Not)A;Brand\";v=\"24\", \"Google Chrome\";v=\"116\"",
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": "\"Windows\"",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36"
}

coo = {
    'TealeafAkaSid': '_BQ57vSnTWetihQK4TQNdHOsQ4SxCJfY',
    'visitorId': '018B480FDCC1020182CE06752C5DEF66',
    'sapphire': '1',
    '_gcl_au': '1.1.979185579.1697721173',
    'crl8.fpcuid': 'b606ef6d-7f89-4511-8f42-f850d9dfde6b',
    'UserLocation': '60430|41.55542439949306|-87.66399297449533|IL|US',
    'fiatsCookie': 'DSI_1460|DSN_Homewood|DSZ_60430',
    'sddStore': 'DSI_1460|DSN_Homewood|DSZ_60430',
    'ci_pixmgr': 'other',
    # 'egsSessionId': '741bb6e1-fdbf-416f-ac63-938846de7a51',
    # 'accessToken': 'eyJraWQiOiJlYXMyIiwiYWxnIjoiUlMyNTYifQ.eyJzdWIiOiI4ZmJmY2Q1Yy0xMTAyLTRjMGEtYjhiNC02NzNhOGNjZjJiMzciLCJpc3MiOiJNSTYiLCJleHAiOjE3MDE0MjU5MDIsImlhdCI6MTcwMTMzOTUwMiwianRpIjoiVEdULjc2MGIwNGQ5ZWJkMjQ1ZWZhOWM1ZTdmNDRiNmY5ZjYwLWwiLCJza3kiOiJlYXMyIiwic3V0IjoiRyIsImRpZCI6ImEwMjdmNWZhYTQ2NDgwMjRhOWJiZDIzYjY3YjFiMTRmYjgzZmIwMTY3Y2JlYTI1YjJiZGExYjc2NzJlYWNmMjEiLCJzY28iOiJlY29tLm5vbmUsb3BlbmlkIiwiY2xpIjoiZWNvbS13ZWItMS4wLjAiLCJhc2wiOiJMIn0.gEL0bCKsH_lBvrdxV0ycqdVDlXenG1awqzx6dTSgUpK8tL7zVhRmI5JzTVBQc041dz_okXkRXEJCwtueMmAsPn776lumuGuh-1HjKswfNu11GCn9C5iL3I0YUWEqw2x9X4U7ZhMBoTPFQP-7IMISlcOOQIiKprvqwgpBbW6-qzRiEU5xrsY1k1MUhKTPO6p26139pzQgyhJ7HAm_0X_q8PhLoOzgAZ921FZFqQv8TNKhXjGCvFF6v2DbZQwPJRm4kieQwZDizSMqYgnYn5CFNwHl8o9CBX_l58DgKEUIZglQR6rIrTL2AwH1GyYcXnKvBwBfaJX3R8zs50XZtgAGVg',
    # 'idToken': 'eyJhbGciOiJub25lIn0.eyJzdWIiOiI4ZmJmY2Q1Yy0xMTAyLTRjMGEtYjhiNC02NzNhOGNjZjJiMzciLCJpc3MiOiJNSTYiLCJleHAiOjE3MDE0MjU5MDIsImlhdCI6MTcwMTMzOTUwMiwiYXNzIjoiTCIsInN1dCI6IkciLCJjbGkiOiJlY29tLXdlYi0xLjAuMCIsInBybyI6eyJmbiI6bnVsbCwiZW0iOm51bGwsInBoIjpmYWxzZSwibGVkIjpudWxsLCJsdHkiOmZhbHNlfX0.',
    # 'refreshToken': 'qJBCfB1afQ_xQTjf0G7cV9SmS9ft_AUQC_Daq7m-1n9BeZjDKpWq1m958EBJh9IDlrVIVreologv-IcwTgrGJA',
    '__gads': 'ID=04521ce72c22e032:T=1697721085:RT=1701339503:S=ALNI_Mb_eM0GqRVI4YbevUQiLLIOHMGCFw',
    '__gpi': 'UID=00000d9af0ee6fb1:T=1697721085:RT=1701339503:S=ALNI_Mb8focbaFTYFdc-33NL0iYm9ZxusA',
    'ffsession': '{%22sessionHash%22:%22aa1bac48dc9831701169204050%22%2C%22prevPageName%22:%22grocery:%20frozen%20foods:%20frozen%20breakfast%20food%22%2C%22prevPageType%22:%22level%203%22%2C%22prevPageUrl%22:%22https://www.target.com/c/frozen-breakfast-food-foods-grocery/-/N-5xsza%22%2C%22sessionHit%22:10%2C%22prevSearchTerm%22:%22non-search%22}',
    '_mitata': 'ZGMwNTJiMzhiNWMyMWE5N2UxZjgwNzMzYzYxNGRlOTZjZjNlNGFiMjcwMzc2MzIwODU0ZDAzZDI4MzJiYjQzOQ==_/@#/1701339917_/@#/czrmMynYzv6QyCM3_/@#/ZDYxODRiYjhiYjdmMzEyMjhmNzVhMjBlZWFjNDAyZDJhZDlhOTNlYzY4Yzc1MGJiOTMxOTU4N2JkZTIwZTUyNA==_/@#/000',
}
def cat_links(pincode):
    dataset = []
    driver = webdriver.Chrome(service=ChromeService(
        ChromeDriverManager().install()))
    driver.get('https://www.target.com/c/grocery/-/N-5xt1a')
    input(f'Change pincode to {pincode} and press enter: ')
    # time.sleep(10)
    # driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div[1]/div/div[1]/div[1]/button/div/span').click()
    # time.sleep(5)
    # driver.find_element(By.XPATH,'/html/body/div[4]/div/div/div[2]/div/div[1]/div/div[1]/div/input').clear()
    # time.sleep(5)
    # driver.find_element(By.ID,'zip-code').send_keys(str(pincode))
    # time.sleep(5)
    # driver.find_element(By.XPATH,'/html/body/div[4]/div/div/div[2]/div/div[2]/button').click()
    # time.sleep(5)
    try:
        driver.find_element(
            By.XPATH, '/html/body/div[1]/div[2]/main/div/div[1]/div/div[2]/div/div/button').click()
    except:
        pass
    time.sleep(5)
    s = BeautifulSoup(driver.page_source, "html.parser")
    links = s.select_one(
        ".h-flex-direction-row.h-display-flex.h-flex-wrap.h-flex-justify-center").find_all("li")[1:]
    for i in links:
        driver.get("https://www.target.com" + i.find("a")["href"])
        try:
            time.sleep(5)
            driver.find_element(
                By.XPATH, '//*[@id="pageBodyContainer"]/div/div[1]/div/div[2]/div/div/button').click()
            time.sleep(3)
        except:
            pass
        s1 = BeautifulSoup(driver.page_source, "html.parser")
        try:
            gl = s1.find("ul").find_all("li")
        except:
            gl = []
        for q in gl:
            try:
                link = q.find("a")["href"]
            except:
                link = None
            if link:
                dataset.append([i.text, q.text, link])
    df = pd.DataFrame(dataset)
    df.to_excel(f"category_links_target.xlsx")
    driver.quit()


def product_links(pro_li):
    url = "https://redsky.target.com/redsky_aggregations/v1/web/plp_search_v2"
    headers = {
        'authority': 'redsky.target.com',
        'accept': 'application/json',
        'accept-language': 'en-US,en;q=0.9',
        'cookie': 'TealeafAkaSid=MgMjEvN2FU2Vkn6RjDYRXv1GLpwRJ6dP; visitorId=018B4718A59302019EC544D7283F1E05; sapphire=1; ci_pixmgr=other; _gcl_au=1.1.960206792.1697705207; crl8.fpcuid=74492362-20c1-463a-b7c8-a7b3c8b7ea8f; UserLocation=60630|41.965761954322524|-87.76362898353945|IL|US; fiatsCookie=DSI_3307|DSN_Chicago%20Wicker%20Park|DSZ_60622; _gid=GA1.2.1183860157.1699168442; _ce.clock_event=1; _ce.clock_data=-931%2C223.233.121.161%2C1%2Cf529a32073a22388a8370c39e9b93c86; _ga=GA1.2.782197828.1699168442; _ga_N15GRPKXME=GS1.1.1699168442.1.1.1699168470.32.0.0; _ce.s=v~5201ee79cc2bbcb3e7c02797510daf3bfcaa558e~lcw~1699168470226~lva~1699168442828~vpv~0~v11.fhb~1699168443768~v11.lhb~1699168449039~v11.cs~316943~v11.s~e68d22d0-7baa-11ee-9fe2-fb73ddc9c3f0~v11.sla~1699168470573~v11.send~1699168453916~lcw~1699168470573; ffsession={%22sessionHash%22:%22f83561def6efe1699245537029%22}; __gads=ID=0e70dc00619d5d00:T=1697704892:RT=1699245538:S=ALNI_MbCWts2qdevbDmidCpzranB7PuKXA; __gpi=UID=00000c6759d3a73e:T=1697704892:RT=1699245538:S=ALNI_MaiO-BZig0n2q9zCkN0YnkwQ8_cJg; sddStore=DSI_3314|DSN_Chicago%20Mayfair|DSZ_60630; egsSessionId=3a83eef0-7537-4609-b0f2-d5878a78230c; accessToken=eyJraWQiOiJlYXMyIiwiYWxnIjoiUlMyNTYifQ.eyJzdWIiOiIxNWM2YzFhMC02OTdlLTQ3MmQtYTdkYS05NWM5ZjA5OWJiMjEiLCJpc3MiOiJNSTYiLCJleHAiOjE2OTkzMzE5NDAsImlhdCI6MTY5OTI0NTU0MCwianRpIjoiVEdULjljMjY4MzYxNTBiZTRlZTViNjQ3Y2Q5ZTFjZDBkM2JlLWwiLCJza3kiOiJlYXMyIiwic3V0IjoiRyIsImRpZCI6IjgxMGE3ZmNmNzVkMzIxZmZjYzExOTk3MDk4N2Y0OTg2Yjc0ZTM2ZTA1ZjI4NjRmOGFiOWQzMDI1YmFjMTgyMGQiLCJzY28iOiJlY29tLm5vbmUsb3BlbmlkIiwiY2xpIjoiZWNvbS13ZWItMS4wLjAiLCJhc2wiOiJMIn0.UskB45OMB1AatRtPwEZx9Pluul41IXGbDUyfz67PNVL1t3ejWOn38kflY-_QXgZfcoavAvdmub3A9wWi-zaZ9D8-aKQ5mU3RfX9d57FiegOjn9rmPzVtFbt1cS6IRLjYwzUcesriLLvr5Y464ssdnmKuku3m5pf9nEij-4Q2tnYF_tf7W51MlzHCL4O1EoyKcxf-fcAl-Drb9GgMNVFUiZ16gbgu7UXlHV-8jy2vJEYRqH4ZXuy2pl3JWUy3YRobqajZ8TceDTpafYB-BBW5DhN-PNCFASFBsWxevJcjlJYIyG1DvO1b07yfeoFQ2NZ0P3wQlYixXG7g0MDogRohQg; idToken=eyJhbGciOiJub25lIn0.eyJzdWIiOiIxNWM2YzFhMC02OTdlLTQ3MmQtYTdkYS05NWM5ZjA5OWJiMjEiLCJpc3MiOiJNSTYiLCJleHAiOjE2OTkzMzE5NDAsImlhdCI6MTY5OTI0NTU0MCwiYXNzIjoiTCIsInN1dCI6IkciLCJjbGkiOiJlY29tLXdlYi0xLjAuMCIsInBybyI6eyJmbiI6bnVsbCwiZW0iOm51bGwsInBoIjpmYWxzZSwibGVkIjpudWxsLCJsdHkiOmZhbHNlfX0.; refreshToken=j4Zfy_hOUnDSqjypVSeBrkyzHRFpFuMepu25gyi434HZKqGoF-XoBgG8BKsA4Jtfucy3ZmUwlorliXFaYp6sJQ; _mitata=MmRmZmMxMTA5ZTY0MTAwZTU3NzkzYWUxODUyZGQ5YWFhZDhmMTc5NTYzZDMyYWM0NDkzZWExYzc3ODk4NGI4OA==_/@#/1699245804_/@#/cQfGMCd1Q7e6X2Ll_/@#/ZjY3NjYxODY5NzA0MzE3ZTc3ZDNmMDZiNGRmNTBjNTllYThkMGNhMzQ0M2NlZTg1OTc0ZTkwYzkwMTgyNGJlNA==_/@#/000; TealeafAkaSid=MgMjEvN2FU2Vkn6RjDYRXv1GLpwRJ6dP; sapphire=1; _mitata=MmQ3ZDlmMDU3MGU4MjFmY2I0NjYwMWVjMjMyOWNlM2JhNWMwMmFiOGQ2MjQwZDczNGQ0MDgyNjY2ZWQ2NjhmMQ==_/@#/1699246051_/@#/cQfGMCd1Q7e6X2Ll_/@#/M2MwZWU0YjU5MDg5NThhN2VlM2MzZmM4Njc1NWQyZGI1Y2ViYThlYzYzYjY3ZGY1Yzg1ZDQ5ZWM0NjJhNzM0OA==_/@#/000',
        'origin': 'https://www.target.com',
        'referer': 'https://www.target.com/c/fresh-vegetables-produce-grocery/-/N-4tglh',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'
    }

    dataset = []
    session = tls_client.Session(client_identifier="chrome112",random_tls_extension_order=True)
    workbook = load_workbook(filename=r"category_links_target.xlsx")
    worksheet = workbook.active
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        if row[3]:
            cat = row[3].split("?")[0].split("/")[-1].split("-")[-1]
            print(idx, '--------------', cat)
            # v_id = generate_random_hex_string(32)
            v_id = '018B4718A59302019EC544D7283F1E05'
            for i in range(0, 18000, 24):
                print(i)
                url = f"https://redsky.target.com/redsky_aggregations/v1/web/plp_search_v2?key=9f36aeafbe60771e321a7cc95a78140772ab3e96&category={cat}&channel=WEB&count=24&default_purchasability_filter=true&include_sponsored=true&new_search=false&offset={i}&page=%2Fc%2F{cat}&platform=desktop&pricing_store_id=2799&scheduled_delivery_store_id=1889&store_ids=2799%2C1889%2C3207%2C3402%2C2781&useragent=Mozilla%2F5.0+%28Windows+NT+10.0%3B+Win64%3B+x64%29+AppleWebKit%2F537.36+%28KHTML%2C+like+Gecko%29+Chrome%2F118.0.0.0+Safari%2F537.36&faceted_value=akkosZ5zl7w&visitor_id=018B4718A59302019EC544D7283F1E05&zip=60603"
                # params = {
                #     "key": "9f36aeafbe60771e321a7cc95a78140772ab3e96",
                #     "category": cat,
                #     "channel": "WEB",
                #     "count": 24,
                #     "default_purchasability_filter": True,
                #     "include_sponsored": True,
                #     "new_search": False,
                #     "offset": i,
                #     "page": f"/c/{cat}",
                #     "platform": "desktop",
                #     "pricing_store_id": 3385,
                #     "scheduled_delivery_store_id": 836,
                #     "store_ids": "3385,3268,3277,3334,3321",
                #     "useragent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
                #     "visitor_id": v_id,
                #     "zip": 60139
                # }
                retry = 5
                while True:
                    try:
                        # proxy_url = random.choice(pro_li)
                        # prox = {
                        #     'http': f'http://{proxy_url}',
                        #     'https': f'http://{proxy_url}',
                        # }
                        # response = requests.get(
                        #     url, headers=headers,proxies=prox).json()
                        response = session.get(
                            url=url,
                            headers=headers,
                            cookies=cookies,
                           
                           
                            proxy='http://geonode_SxA5BnHLFX-country-HU:6aaed733-56f0-4ffc-9b44-b7cc64c5ca87@rotating-residential.geonode.com:9003').json()
                        break
                    except:
                        retry -= 1
                        time.sleep(2)
                        if retry == 0:
                            print('failed')
                            break
                time.sleep(3)        
                try:
                    if not response['data']['search']['products']:
                        break
                except:
                    break
                for pr in response['data']['search']['products']:
                    dataset.append(
                        [pr['tcin'], pr['item']['enrichment']['buy_url']])
    df = pd.DataFrame(dataset, columns=['tcin', 'url'])
    df.drop_duplicates(subset=['url'], inplace=True)
    df.to_excel("product_links_60603.xlsx", index=False)


def product_cat():
    headers = {
        'authority': 'www.target.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-US,en;q=0.9',
        'cookie': 'TealeafAkaSid=MgMjEvN2FU2Vkn6RjDYRXv1GLpwRJ6dP; visitorId=018B4718A59302019EC544D7283F1E05; sapphire=1; mdLogger=false; kampyle_userid=f906-261b-3153-31c9-55c6-6736-9f6d-5ba1; ci_pixmgr=other; _gcl_au=1.1.960206792.1697705207; crl8.fpcuid=74492362-20c1-463a-b7c8-a7b3c8b7ea8f; kampylePageLoadedTimestamp=1698035605479; LAST_INVITATION_VIEW=1698035619800; salsify_session_id=d31b6554-8ad0-47d1-9f83-0636f1974049; fiatsCookie=DSI_3307|DSN_Chicago%20Wicker%20Park|DSZ_60622; _gid=GA1.2.1183860157.1699168442; _ce.clock_event=1; _ce.clock_data=-931%2C223.233.121.161%2C1%2Cf529a32073a22388a8370c39e9b93c86; _ga=GA1.2.782197828.1699168442; _ga_N15GRPKXME=GS1.1.1699168442.1.1.1699168470.32.0.0; _ce.s=v~5201ee79cc2bbcb3e7c02797510daf3bfcaa558e~lcw~1699168470226~lva~1699168442828~vpv~0~v11.fhb~1699168443768~v11.lhb~1699168449039~v11.cs~316943~v11.s~e68d22d0-7baa-11ee-9fe2-fb73ddc9c3f0~v11.sla~1699168470573~v11.send~1699168453916~lcw~1699168470573; GuestLocation=334401|28.010|73.320|RJ|IN; ffsession={%22sessionHash%22:%22bb948850700b51699253253970%22}; egsSessionId=d1660bb9-468a-47b6-9f26-032f2de3b9c2; accessToken=eyJraWQiOiJlYXMyIiwiYWxnIjoiUlMyNTYifQ.eyJzdWIiOiIxNWM2YzFhMC02OTdlLTQ3MmQtYTdkYS05NWM5ZjA5OWJiMjEiLCJpc3MiOiJNSTYiLCJleHAiOjE2OTkzMzk2NTYsImlhdCI6MTY5OTI1MzI1NiwianRpIjoiVEdULmRmNTA4N2VhN2I4NjRhNzZiOGI1MDMwMDJhYTNhN2MwLWwiLCJza3kiOiJlYXMyIiwic3V0IjoiRyIsImRpZCI6IjgxMGE3ZmNmNzVkMzIxZmZjYzExOTk3MDk4N2Y0OTg2Yjc0ZTM2ZTA1ZjI4NjRmOGFiOWQzMDI1YmFjMTgyMGQiLCJzY28iOiJlY29tLm5vbmUsb3BlbmlkIiwiY2xpIjoiZWNvbS13ZWItMS4wLjAiLCJhc2wiOiJMIn0.wZBngWI3PpnG_tT3GvC94rtGRWn-ePpqPv_1i1RaAU2kIBdvqaszy4Qx9GhLmMOqmuUprFT30Uad_39MkWPToFajLnL424AbnVQfZPiubHUnX2w9_1xF2ad9Lt6AeoYoJTDOl6U5RtaL6YNde1AIJqrGoyxW_BaHTJBHPZN-y1YcXce_Df09QZ1GNKurjMesTyAFc7_kTVtKLLT-HCm9T8niaDr_PWbrc3N0Y269JDYiN-9HK2n2xRIfQdHpRWRNv3vs46W6fk3Q9u3mIvGZuC4jyxJ_n_z4oBwU_TgjR1R4KMPGTM0hExdLpNPIR33SDRpMraPtzviZ9F39HipBEA; idToken=eyJhbGciOiJub25lIn0.eyJzdWIiOiIxNWM2YzFhMC02OTdlLTQ3MmQtYTdkYS05NWM5ZjA5OWJiMjEiLCJpc3MiOiJNSTYiLCJleHAiOjE2OTkzMzk2NTYsImlhdCI6MTY5OTI1MzI1NiwiYXNzIjoiTCIsInN1dCI6IkciLCJjbGkiOiJlY29tLXdlYi0xLjAuMCIsInBybyI6eyJmbiI6bnVsbCwiZW0iOm51bGwsInBoIjpmYWxzZSwibGVkIjpudWxsLCJsdHkiOmZhbHNlfX0.; refreshToken=vVnjwwQ1bsgNxi-V7XctSqfQUbBOfmDN2cHgtPDUsvPTHPoxmsguXuMKjP4EWVbblQuJujw1cAGTXesZt9stiA; UserLocation=60603|41.8808797036521|-87.62812146839536|IL|US; sddStore=DSI_1889|DSN_Chicago%20South%20Loop|DSZ_60603; __gads=ID=0e70dc00619d5d00:T=1697704892:RT=1699253581:S=ALNI_MbCWts2qdevbDmidCpzranB7PuKXA; __gpi=UID=00000c6759d3a73e:T=1697704892:RT=1699253581:S=ALNI_MaiO-BZig0n2q9zCkN0YnkwQ8_cJg; dteRfWys=OnIBroeU; kampyleUserSession=1699253606007; kampyleUserSessionsCount=55; kampyleSessionPageCounter=1; kampyleUserPercentile=93.92286095788032; TealeafAkaSid=MgMjEvN2FU2Vkn6RjDYRXv1GLpwRJ6dP; sapphire=1',
        #   'if-none-match': '"f3rc9t10dx5hdi"',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'
    }
    d = []
    workbook = load_workbook(filename=r"product_links_60603 1.xlsx")
    worksheet = workbook.active

    def scrap(row):
        print(row[0])
    # for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        r = requests.get(row[1], headers=headers)
        s = BeautifulSoup(r.text, "html.parser")
        links = s.select(".styles__LinkContainer-sc-1cq6bg4-0.iUWtvr")
        word = links[1].text.replace("/", "").strip()
        if word == "Grocery":
            cat1 = links[2].text.replace("/", "").strip()
            cat2 = links[3].text.replace("/", "").strip()
            today = datetime.today().date()
            formatted_today = today.strftime('%Y-%m-%d')
            title = s.select_one('#pdp-product-title-id').text
            img = s.select_one('#PdpImageGallerySection').find('img')['src']
            primary_barcode = ''
            current_retail = ''
            net_weight = ''
            reg_retail = ''
            sc = s.find_all('script')
            for i in sc:
                if 'window.__FLAGS__' in i.text:
                    input_string = i.text.split("'__TGT_DATA__'")[-1].split("'__WEB_CLUSTER__'")[
                        0].replace('\\u003c/B\\u003e', '').strip()
                    # Define the regular expression pattern
                    pattern = r'primary_barcode\\":\\"(\d+)\\"'

                    # Use re.search to find the first match
                    match = re.search(pattern, input_string)

                    # Check if a match is found
                    if match:
                        # Extract the primary barcode from the match object
                        primary_barcode = match.group(1)
                    else:
                        pass
                    # Define the regular expression pattern
                    pattern = r'Net weight:(.*?)\"'

                    # Use re.search to find the first match
                    match = re.search(pattern, input_string)

                    # Check if a match is found
                    if match:
                        # Extract the net weight from the match object
                        net_weight = match.group(1).replace('\\', '')

                    else:
                        pass

                    # Define the regular expression pattern
                    pattern = r'reg_retail\\":([^,]+)'

                    # Use re.search to find the first match
                    match = re.search(pattern, input_string)

                    # Check if a match is found
                    if match:
                        # Extract the net weight from the match object
                        reg_retail = match.group(1).replace('\\', '')

                    else:
                        pass

                    # Define the regular expression pattern
                    pattern = r'current_retail\\":([^,]+)'

                    # Use re.search to find the first match
                    match = re.search(pattern, input_string)

                    # Check if a match is found
                    if match:
                        # Extract the net weight from the match object
                        current_retail = match.group(1).replace('\\', '')

                    else:
                        pass
            if reg_retail == current_retail:
                current_retail = ''
            
            if net_weight == '':
                matched_items = find_matches_reverse(title)
                if matched_items:
                    for match in matched_items:
                        try:
                            net_weight = match[1][::-1] + ' ' + match[0][::-1]
                        except:
                            pass     
            data = {
                "zipcode": "60603",
                "store_name": "Chicago State St.",
                "store_location": "1 S State Street, Chicago, IL",
                "store_logo": "https://assets.targetimg1.com/webui/store-locator/targetlogo-6.jpeg",
                "category": cat1,
                "sub_category": cat2.replace("\u200e", ""),
                "product_title": title,
                "weight": net_weight,
                "regular_price": reg_retail,
                "sale_price":  current_retail,
                "image_url": img,
                "url": row[1],
                "upc": trim_and_pad(primary_barcode),
                "crawl_date": formatted_today}
            print(data)
            d.append(data)
    with concurrent.futures.ThreadPoolExecutor(8) as ex:
        ex.map(scrap, worksheet.iter_rows(min_row=2, values_only=True))

    df = pd.DataFrame(d)
    df.to_excel("Target_full_60603.xlsx")


# cat_links(60603)
# product_links()
# product_cat()

def scrapper_60603(proxy_data):
    product_links(proxy_data)
    product_cat()
