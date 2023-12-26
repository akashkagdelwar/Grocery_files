import random
import requests
from bs4 import BeautifulSoup
import pandas as pd
import pyautogui
from openpyxl import load_workbook
import json
# import undetected_chromedriver as uc
import pickle
import time
import urllib3
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
# import undetected_chromedriver as uc
from datetime import datetime
# import undetected_chromedriver as uc
# from seleniumwire import webdriver
import tls_client
from selenium.webdriver.firefox.service import Service as FirefoxService
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import re
from psycopg2.extras import execute_values
import psycopg2


query = ''' Select * from walmart_products'''            
def get_pg():
    conn = psycopg2.connect(
        host="db-postgresql-sfo3-smartbot-1db-do-user-8157534-0.b.db.ondigitalocean.com",
        port="25061",
        database="alpha_pool_2023",
        user="doadmin",
        password="mq2i4pwpvlen6mho",
      
    )
    cursor = conn.cursor()
    return cursor, conn

def batch_insert(dataset, insert_query):
    cursor, connection = get_pg()
    tup = [tuple(data.values()) for data in dataset]
    execute_values(cursor, insert_query, tup)
    connection.commit()
    cursor.close()
    connection.close()
# headers = {
#     'authority': 'www.walmart.com',
#     'method': 'GET',
#     'scheme': 'https',
#     'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
#     'Accept-Encoding': 'gzip, deflate, br',
#     'Accept-Language': 'en-US,en;q=0.9',
#     'Cache-Control': 'max-age=0',
#     'Referer': 'https://www.walmart.com/cp/food/976759?povid=GlobalNav_rWeb_Grocery_Grocery_ShopAll',
#     'Sec-Ch-Ua': '"Chromium";v="116", "Not)A;Brand";v="24", "Google Chrome";v="116"',
#     'Sec-Ch-Ua-Mobile': '?0',
#     'Sec-Ch-Ua-Platform': '"Windows"',
#     'Sec-Fetch-Dest': 'document',
#     'Sec-Fetch-Mode': 'navigate',
#     'Sec-Fetch-Site': 'same-origin',
#     'Sec-Fetch-User': '?1',
#     'Upgrade-Insecure-Requests': '1',
#     'User-Agent': ua.chrome}

# Disable all warnings from urllib3


def pro():
    proxy_data = [
        "residential.proxyomega.com:10000:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10001:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10002:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10003:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac",
        "residential.proxyomega.com:10004:jyothish_1822:e9700d-ec46f9-6317aa-e45cc5-34e1ac"
    ]

    # Create a list of tuples from the data
    proxies = [tuple(proxy.split(':')) for proxy in proxy_data]

    # Randomly select a proxy from the list
    random_proxy = random.choice(proxies)
    proxy_url = f"http://{random_proxy[2]}:{random_proxy[3]}@{random_proxy[0]}:{random_proxy[1]}"

    # Set up the proxy dictionary
    proxy = {
        'http': proxy_url,
        'https': proxy_url,
    }
    return proxy


def driver1():
    options = webdriver.EdgeOptions()
    options.use_chromium = True
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    # options.add_argument("--headless")
    options.add_argument("--disable-blink-features")
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-speech-api")
    options.add_argument("--disable-web-security")
    options.add_argument("--start-maximized")
    options.add_argument("--window-size=1920,1080")
    # options.add_experimental_option(
    #      "prefs", {'profile.managed_default_content_settings.javascript': 2})
    # driver = uc.Chrome(options=options)
    driver = uc.Chrome(use_subprocess=True)
    driver.implicitly_wait(10)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver

def trim_and_pad(number):
    # Convert the number to a string and remove the last character
    trimmed_number = str(number)[:-1]
    # Calculate how many zeros to add at the start
    zeros_to_add = 13 - len(trimmed_number)
    # Add the required number of zeros at the start
    modified_number = str('0' * zeros_to_add) + str(trimmed_number)
    return str(modified_number)


insert_query = """INSERT INTO walmart_products ("product_id", "upc", "url")
                  VALUES %s
                  ON CONFLICT DO NOTHING
               """

pyautogui.FAILSAFE = False
cookies = {
    # 'vtc': 'TagNVl2Yx5CTeiVNItZZ8w',
    'adblocked': 'false',
    'ACID': 'a84881ac-854a-410e-9a86-aeaa4813e2db',
    'hasACID': 'true',
    # '_m': '9',
    # 'hasLocData': '1',
    # 'abqme': 'true',
    # '_pxhd': '772835dbd0225cecd466b423af289c6056f600d32cf868845d9fe01bb3cf103e:6683cf87-8cfc-11ee-8eb8-f39397f3fb4d',
    # 'TBV': 'f',
    # 'pxcts': '6698e8a1-8cfc-11ee-a2c9-9b8989c7714c',
    # '_pxvid': '6698d8de-8cfc-11ee-a2c9-f95397fabcbd',
    # 'userAppVersion': 'us-web-1.104.4-c58d431eba530d8e7fb2d4a90a93a0a51ed5a9ef-1113',
    # '_astc': '4566d454a948ad885257e9f043948c93',
    # 'bstc': 'X8e_-0QVzKYaRFNFbH8Egk',
    # 'mobileweb': '0',
    # 'xpth': 'x-o-mart%2BB2C~x-o-mverified%2Bfalse',
    # 'xpa': '',
    # 'ak_bmsc': '2440E5BE531880E283F699F8D73C63FC~000000000000000000000000000000~YAAQTmM4F+dCyuKLAQAAJL7OFhXXfywOtyyvxwtBzz8mUxTozRGdRs1OeljtpuFZg1J/Wkp318Tt3HeGL8B2tSO/regfAo/cRBgaQOxTpCGl1KLF+LK3kLO8QQesxeq4N5uGLxF8ZmJRFjM7i5jomjyTIWLyJUT84X5WQUApKZ+l/P2JXUG4zyXSs9EK0sIqfsrL6t0groZXLj/krZeMmInM5cMFH4QOfaut5X9eBziBkcr2aOJ7/1r4ya1f8ldaAVf1H2a7i9N8VsAObLqni1o+sNNtbwSd8/1pU7aWjiLccAba7dee28RL6J6jTCowPm2qscXhQhPguHyTIAaCSGuxa1Q024h44Ulwu+ZFpoyNjC2REbqaR/c9Y4BmopCVxTSKBDV9Ym1Wdtr2GA==',
    # 'auth': 'MTAyOTYyMDE4Jje6bs7cK96wyDqL9mRw9Si%2BHTKSTwIBW9KVLxuBuiyz4LHIkc9qAFCpKHoG%2Fi88sdgxyQD4sKWBVQc4omsJ%2Bz8QdG%2F0DqGzK%2Fj%2BSvITEIqO8%2BgWPmwn%2Fyx1o2RajHeU767wuZloTfhm7Wk2Kcjygv699%2F6tFVwuL3qJB39WKV%2B%2FJ%2BMbjpi70jaN6DngU9hCUPkBh3yI2dk39IFxK4k8LjSRXkT9oNSyzUwLUNVMq3oUMk70P8glgOEpLOprhDfMDCcb9mgycy9jtT1uIyOBHY%2F8cc%2BduLhKdJF01eANJ4EeNnux6HSxeNWo%2FGjC1xM6CabjL0zPGc9EN1VKH6w8acuPT0%2FFaA6aRQoqqIpEadchj6SyIphrJZKPB0XcjWeexAPxTlvQCzKl2lcUtJh5%2B0jyrOXbKKhH072NS%2FW0j%2FU%3D',
    'assortmentStoreId': '1401',
    'locGuestData': 'eyJpbnRlbnQiOiJQSUNLVVAiLCJpbnRlbnRTb3VyY2UiOiJHSUNfRnVsZmlsbG1lbnRfUHJlZmVyZW5jZSIsImlzRXhwbGljaXQiOnRydWUsInN0b3JlSW50ZW50IjoiUElDS1VQIiwibWVyZ2VGbGFnIjp0cnVlLCJpc0RlZmF1bHRlZCI6ZmFsc2UsInBpY2t1cCI6eyJub2RlSWQiOiIxNDAxIiwidGltZXN0YW1wIjoxNzAxMDcyNjE3NTQ5LCJzZWxlY3Rpb25UeXBlIjoiQ1VTVE9NRVJfU0VMRUNURUQiLCJzZWxlY3Rpb25Tb3VyY2UiOiJQaWNrdXAgU3RvcmUgU2VsZWN0b3IifSwic2hpcHBpbmdBZGRyZXNzIjp7InRpbWVzdGFtcCI6MTcwMTA3MjYxNzU0OSwidHlwZSI6InBhcnRpYWwtbG9jYXRpb24iLCJnaWZ0QWRkcmVzcyI6ZmFsc2UsInBvc3RhbENvZGUiOiI2MDU2NCIsImNpdHkiOiJOYXBlcnZpbGxlIiwic3RhdGUiOiJJTCIsImRlbGl2ZXJ5U3RvcmVMaXN0IjpbeyJub2RlSWQiOiIxNDAxIiwidHlwZSI6IkRFTElWRVJZIiwidGltZXN0YW1wIjoxNzAxMTg5NzA2NDgwLCJzZWxlY3Rpb25UeXBlIjoiTFNfU0VMRUNURUQiLCJzZWxlY3Rpb25Tb3VyY2UiOiJaSVBfQ09ERV9CWV9VU0VSIn1dfSwicG9zdGFsQ29kZSI6eyJ0aW1lc3RhbXAiOjE3MDEwNzI2MTc1NDksImJhc2UiOiI2MDU2NCJ9LCJtcCI6W10sInZhbGlkYXRlS2V5IjoicHJvZDp2MjphODQ4ODFhYy04NTRhLTQxMGUtOWE4Ni1hZWFhNDgxM2UyZGIifQ%3D%3D',
    # 'akavpau_p1': '1701190307~id=c0b1450fff772be93dd2705166597815',
    'locDataV3': 'eyJpc0RlZmF1bHRlZCI6ZmFsc2UsImlzRXhwbGljaXQiOnRydWUsImludGVudCI6IlBJQ0tVUCIsInBpY2t1cCI6W3siYnVJZCI6IjAiLCJub2RlSWQiOiIxNDAxIiwiZGlzcGxheU5hbWUiOiJOYXBlcnZpbGxlIFN1cGVyY2VudGVyIiwibm9kZVR5cGUiOiJTVE9SRSIsImFkZHJlc3MiOnsicG9zdGFsQ29kZSI6IjYwNTY0IiwiYWRkcmVzc0xpbmUxIjoiMjU1MiBXIDc1dGggU3QiLCJjaXR5IjoiTmFwZXJ2aWxsZSIsInN0YXRlIjoiSUwiLCJjb3VudHJ5IjoiVVMiLCJwb3N0YWxDb2RlOSI6IjYwNTY0LTc1NzIifSwiZ2VvUG9pbnQiOnsibGF0aXR1ZGUiOjQxLjc0NTI1NSwibG9uZ2l0dWRlIjotODguMTk4MjE1fSwiaXNHbGFzc0VuYWJsZWQiOnRydWUsInNjaGVkdWxlZEVuYWJsZWQiOnRydWUsInVuU2NoZWR1bGVkRW5hYmxlZCI6dHJ1ZSwiaHViTm9kZUlkIjoiMTQwMSIsInN0b3JlSHJzIjoiMDY6MDAtMjM6MDAiLCJzdXBwb3J0ZWRBY2Nlc3NUeXBlcyI6WyJQSUNLVVBfSU5TVE9SRSIsIlBJQ0tVUF9DVVJCU0lERSJdLCJzZWxlY3Rpb25UeXBlIjoiQ1VTVE9NRVJfU0VMRUNURUQifV0sInNoaXBwaW5nQWRkcmVzcyI6eyJsYXRpdHVkZSI6NDEuNzA1OCwibG9uZ2l0dWRlIjotODguMjAwNSwicG9zdGFsQ29kZSI6IjYwNTY0IiwiY2l0eSI6Ik5hcGVydmlsbGUiLCJzdGF0ZSI6IklMIiwiY291bnRyeUNvZGUiOiJVU0EiLCJnaWZ0QWRkcmVzcyI6ZmFsc2UsInRpbWVab25lIjoiQW1lcmljYS9DaGljYWdvIn0sImFzc29ydG1lbnQiOnsibm9kZUlkIjoiMTQwMSIsImRpc3BsYXlOYW1lIjoiTmFwZXJ2aWxsZSBTdXBlcmNlbnRlciIsImludGVudCI6IlBJQ0tVUCJ9LCJpbnRlbnRTb3VyY2UiOiJHSUNfRnVsZmlsbG1lbnRfUHJlZmVyZW5jZSIsImluc3RvcmUiOmZhbHNlLCJkZWxpdmVyeSI6eyJidUlkIjoiMCIsIm5vZGVJZCI6IjE0MDEiLCJkaXNwbGF5TmFtZSI6Ik5hcGVydmlsbGUgU3VwZXJjZW50ZXIiLCJub2RlVHlwZSI6IlNUT1JFIiwiYWRkcmVzcyI6eyJwb3N0YWxDb2RlIjoiNjA1NjQiLCJhZGRyZXNzTGluZTEiOiIyNTUyIFcgNzV0aCBTdCIsImNpdHkiOiJOYXBlcnZpbGxlIiwic3RhdGUiOiJJTCIsImNvdW50cnkiOiJVUyIsInBvc3RhbENvZGU5IjoiNjA1NjQtNzU3MiJ9LCJnZW9Qb2ludCI6eyJsYXRpdHVkZSI6NDEuNzQ1MjU1LCJsb25naXR1ZGUiOi04OC4xOTgyMTV9LCJpc0dsYXNzRW5hYmxlZCI6dHJ1ZSwic2NoZWR1bGVkRW5hYmxlZCI6dHJ1ZSwidW5TY2hlZHVsZWRFbmFibGVkIjp0cnVlLCJhY2Nlc3NQb2ludHMiOlt7ImFjY2Vzc1R5cGUiOiJERUxJVkVSWV9BRERSRVNTIn1dLCJodWJOb2RlSWQiOiIxNDAxIiwiaXNFeHByZXNzRGVsaXZlcnlPbmx5IjpmYWxzZSwic3VwcG9ydGVkQWNjZXNzVHlwZXMiOlsiREVMSVZFUllfQUREUkVTUyJdLCJzZWxlY3Rpb25UeXBlIjoiTFNfU0VMRUNURUQifSwicmVmcmVzaEF0IjoxNzAxMTkzMzA3NTI2LCJ2YWxpZGF0ZUtleSI6InByb2Q6djI6YTg0ODgxYWMtODU0YS00MTBlLTlhODYtYWVhYTQ4MTNlMmRiIn0%3D',
    'xptc': 'assortmentStoreId%2B1401',
    # 'xpm': '3%2B1701189707%2BTagNVl2Yx5CTeiVNItZZ8w~%2B0',
    # 'dimensionData': '843',
    # 'bm_mi': 'D4877FAE81069F4255D3C396918CD5D1~YAAQXGM4F9CKnRGMAQAAKTnPFhW9jU2NnmSutvnreaHoRKJBnuG4jaQ4yhjhUVmgAExlGrlJP27YRYJ9aqd3jeZc527Wu8Cw3vn2gqEcclyHoXbnph/LUKHoj3TotQTkoxgB6/NHllw7MXBPLK5m9iGpLvG5LGjYLKdAXoMfaZegYVTInp3BT8Vwrja+hae9vK1UKjfR6wZ+hqpukcQtgL/OQpo9XouQJ13aGm+4H5/QucyNcsTTfDoo6QSpnKbxAhGey7iNBBJA1YAZ2YxDOAeErtft0o0HRoTXM/Xvy1cWls7Kcqy48C5155V2A8n1uYAS+Y5eiiRsi4KOOFniNgST~1',
    # 'AID': 'wmlspartner%3Dwmtlabs%3Areflectorid%3D22222222220220085369%3Alastupd%3D1701189732990',
    # 'xptwj': 'qq:a42750c631a14e5f3809:3P1hlNDglfKUUUJU9ZNu6ku0pGOEXGzRKYqs0tKYKaySjcAzji6CVV6aG3jKUZyP7Gb40jNs/Z5oCh27NlQwL1LQKHlSVE8zwopBWLvcE2UvhXmw/m9nnVycI3XANxsGO29cjDZOQ89ds9o/Cd4CQMwfgvO/lNbzdZINB/k=',
    # '_px3': '44ce0223b9b12899e8c5dfeb53da1f07de26ccab55e4d4f795edcfc53096f87e:65kRXBs5cbNmOdmupuhKO/KCQuIg+ZvCdnif3kfDQdNy+2nUvjlehodxDtVlV/6OrCfZtuDJVRWE0u225STmew==:1000:OQT7Foh1HmEVjL3jjS1TXOMW0wDBfg11N7lSlC4F9xITyQaVTkHh+fzLU0k62yx/P9Gex4ShEArUDAmZpz/fr5ZBydk2j9T5fxHas7IgVve8SSKs3pwFRh20xUUTNeoB933n7i+kwWp5MJ0R1cjLC7HjW2SlVLMCWodPq/eJhCjeyWUKXoZEkL26/HiGnCwgijw7zb3KMtAT7EL0A55pugMBIgj6ZYPdqjZ1yjW/OMk=',
    # 'com.wm.reflector': '"reflectorid:22222222220220085369@lastupd:1701189737000@firstcreate:1701072618123"',
    # 'xptwg': '816244560:2565DD78E5FB100:5E6AFD1:92F541E4:D417B7D1:89A75CBE:',
    # 'TS012768cf': '01712679e39d161331ab4a1f2d8ef2a8859656135570ee2e74b4e278c90f23d2596cec82593258c97a97c51ae315aeb3a47d5aa8a3',
    # 'TS01a90220': '01712679e39d161331ab4a1f2d8ef2a8859656135570ee2e74b4e278c90f23d2596cec82593258c97a97c51ae315aeb3a47d5aa8a3',
    # 'TS2a5e0c5c027': '08a436d427ab2000a20c43ff3f8da67a34d1170e10c385688cc43a26d1c731f84ea7375282d71d1208aa031eb5113000a7b26ca0be404ace5d7034e218a08339e04a42deda79e2dd59568c067b3ecde25d3a2bbfbcffd3fdbd8a23ba638e8963',
    # 'akavpau_p2': '1701190337~id=4f8a19d1714c90af2d93f26ba8f01923',
    # 'bm_sv': 'F81723CAA5103FD2D2A5054E29140BED~YAAQXmM4F6TgwN+LAQAAIFzPFhX+2pkLhYuUREEhwyXnEn+W76rsRR/fVZxzY/nXIV7wPQwn/dl2qPEN23KirgJRL1g7za86DEZbvBS+/qVz/jkLfrS55VJ1Y47uk94AiJmbGqUl8//PA8Z2QUO8J3RCobw87jN3f/TrTSzmblrOYFrXKadvOIfy72XqSvVdNdia2wtP0aTb86mECETmOGWbssPqO+MEXA0VR4ifKPPnAbWvuHQJ77n1oyI5I5Fyj7E=~1',
    # '_pxde': 'da16e0b2cbef7c0b4da377dc2b9a4785c7d9e680396ede0ad9e0146623812736:eyJ0aW1lc3RhbXAiOjE3MDExODk3Mzg3NTd9',
}
def mouse():
    screen_width, screen_height = pyautogui.size()
    center_x = screen_width // 2
    center_y = screen_height // 2
    pyautogui.moveTo(center_x, center_y - 160)
    time.sleep(10)
    pyautogui.mouseDown()
    time.sleep(60)
    pyautogui.mouseUp()
    time.sleep(5)


headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
    "Cache-Control": "max-age=0",
}

# Define the list of items to check
items_to_check = ['crv', 'oz', 'fz', 'count', 'ounce', 'fl',
                  'gallon', 'ct', 'lb', 'liter', 'lt', 'pound', 'kg', 'ml']

items_to_check = [item[::-1] for item in items_to_check]

print(items_to_check)

# Create a regex pattern to match any number followed by an element from the list in reverse
pattern = r'(' + '|'.join(re.escape(item)
                          for item in items_to_check) + r')\s*(\d+(?:\.\d+)?)'
# Function to find all matches in a reversed string


def find_matches_reverse(text):
    reversed_text = text[::-1]  # Reverse the input string
    matches = re.findall(pattern, reversed_text, re.IGNORECASE)
    return matches


def cat_links():
    dataset = []
    driver = driver1()
    driver.get('https://www.walmart.com/cp/food/976759',)
    s = BeautifulSoup(driver.page_source, "html.parser")
    cat = s.select("#LeftHandNavList")[1].find_all("li", class_=False)
    for i in cat:
        a_tag = [ro.find('a')['href'].split("?")[0]
                 for ro in i.find_all('li')[1:]]
        for i in a_tag:
            if "cp" in i:
                if "https" not in i:
                    i = "https://www.walmart.com" + i
                driver.get(i)
                s1 = BeautifulSoup(driver.page_source, "html.parser")
                test = s1.find_all("section")
                for q in test:
                    if "Shop by" in q.text.strip():
                        a1 = q.find_all("a")
                        for z in a1:
                            data = z['href'].split("?")[0]
                            if "https" not in data:
                                data = "https://www.walmart.com" + data
                            dataset.append(data)
            else:
                dataset.append(i)

    dataset = [te for te in dataset if "cp" not in te]
    df = pd.DataFrame(dataset, columns=["url"])
    df.drop_duplicates(subset=["url"], inplace=True)
    df.to_excel("category_links.xlsx", index=False)
    driver.quit()


def p_links():
    # Set the proxy URL with credentials
    proxy_url = "http://geonode_SxA5BnHLFX:6aaed733-56f0-4ffc-9b44-b7cc64c5ca87@rotating-residential.geonode.com:9000"
    dataset = []
    count_li = []
    upc_dict = {}
    cursor, connection = get_pg()
    cursor.execute(query)
    # Fetch all the results
    all_results = cursor.fetchall()
    # Don't forget to close the cursor when you're done
    cursor.close()
    for results in all_results:
        upc_dict[str(results[1])] = str(results[2])
    workbook = load_workbook(filename=r"category_links.xlsx")
    worksheet = workbook.active
    session = tls_client.Session(
        client_identifier="safari_ios_15_6", random_tls_extension_order=True)
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        with open("data.pkl", 'wb') as file:
            pickle.dump(dataset, file)
        print(idx, "------------->>>", row[0])
        count = 0
        for pg in range(1, 26):
            print(pg)
            try:
                res = session.get(
                    row[0]+f"?page={pg}&affinityOverride=default",
                    headers=headers,
                    cookies=cookies,
                    allow_redirects=True,
                    # proxy="http://c5f37321e0ed190c06d01b65eda7c210a1e1cfa9:premium_proxy=true@proxy.zenrows.com:8001",)
                    # insecure_skip_verify=True)
                )
            except:
                pass
            time.sleep(4)
            s = BeautifulSoup(res.text, "lxml")
         
            if "Robot or human?" in res.text.strip():
                input('Solve captacha and press enter: ')
            tag = s.find_all('script')
            j = None
            for tg in tag:
                if "initialData" in tg.text.strip():
                    j = json.loads(tg.text)
                    # with open("product_wallmart.josn", "w") as outfile:
                    #     json.dump(j, outfile)
            if j:
                title = j["props"]["pageProps"]["initialData"]["searchResult"]["itemStacks"][0]["title"]
                # count_new = j["props"]["pageProps"]["initialData"]["searchResult"]["itemStacks"][0]["count"]
                # print([title, count_new])
                # count_li.append([title, count_new])
                li = j["props"]["pageProps"]["initialData"]["searchResult"]["itemStacks"][0]["items"]
                pg_countint = int(
                    j["props"]["pageProps"]["initialData"]["searchResult"]["aggregatedCount"])
                iter = int(pg*40)
                if iter > pg_countint:
                    count += 1
                    if count == 2:
                        break
                for pr in li:
                    # title = pr["name"]
                    # image = "https://i5.walmartimages.com/seo/" + pr["imageInfo"]["name"]
                    try:
                        p_url = "https://www.walmart.com" + pr["canonicalUrl"]
                        list_price = pr["priceInfo"]["linePrice"]
                        sale_price = pr["priceInfo"]["wasPrice"]
                        name = pr["name"]
                        img = pr["imageInfo"]["thumbnailUrl"].split('?')[0]
                        today = datetime.today().date()
                        formatted_today = today.strftime('%Y-%m-%d')
                        cat1 = s.select('.w_MSFl')[1].text
                        try:
                            cat2 = s.select('.w_MSFl')[2].text
                        except:
                            cat2 = ''
                        weight = ''
                        matched_items = find_matches_reverse(name)
                        if matched_items:
                            for match in matched_items:
                                try:
                                    weight = match[1][::-1] + \
                                        ' ' + match[0][::-1]
                                except:
                                    pass
                        if sale_price == list_price:
                            sale_price = ''
                        if sale_price:
                            list_price,sale_price = sale_price,list_price   
                        try:
                            upc =  upc_dict[p_url.split('?')[0].split('/')[-1]]
                        except:
                            upc = ''
                        if not title:
                            title = cat2 

                        data_dict = {
                            "zipcode": "60564",
                            "store_name": "Naperville Supercenter",
                            "store_location": "2552 W 75th St, Naperville, IL 60564",
                            "store_logo": "https://i5.walmartimages.com/dfw/63fd9f59-b3e1/7a569e53-f29a-4c3d-bfaf-6f7a158bfadd/v1/walmartLogo.svg",
                            "category": cat1,
                            "sub_category": title,
                            # 'cat3': cat2,
                            "product_title": name,
                            "weight": weight,
                            "regular_price":list_price.replace('$', ''),
                            "sale_price": sale_price.replace('$', '') ,
                            "image_url": img,
                            "url": p_url,
                            "upc": upc,
                            "crawl_date": formatted_today}
                        print(data_dict)
                        dataset.append(data_dict)
                    except Exception as e:
                        pass
            else:
                print('missing')
    df = pd.DataFrame(dataset)
    df.drop_duplicates(subset=["url"], inplace=True)
    writer = pd.ExcelWriter(
        f'OUTPUT - Product links 60564', engine='xlsxwriter', engine_kwargs={"options": {'strings_to_urls': False}})
    df.to_excel(writer, index=False)
    writer.close()


def p_data():
    dataset = []
    missing_dataset = []
    session = tls_client.Session(
        client_identifier="chrome112", random_tls_extension_order=True)
    workbook = load_workbook(filename=r"OUTPUT - Product links 60564.xlsx")
    worksheet = workbook.active
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        with open("missing_upc.pkl", 'wb') as file:
            pickle.dump(dataset, file)
        if not row[12]:    
            print(idx,"------------->>>", row[11])
            try:
                res = session.get(
                        row[11],
                        headers=headers,
                        cookies=cookies,
                        # proxy="http://c5f37321e0ed190c06d01b65eda7c210a1e1cfa9:premium_proxy=true@proxy.zenrows.com:8001",
                        # insecure_skip_verify=True
                    )
            except:
                pass
            time.sleep(4)
            s = BeautifulSoup(res.text, "html.parser")
            if "Robot or human?" in res.text.strip():
                input('Solve captacha and press enter: ')
            tag = s.find_all('script')
            j = None
            try:
                for tg in tag:
                    if "initialData" in tg.text.strip():
                        j = json.loads(tg.text)
                        # with open("product_wallmart.josn", "w") as outfile:
                        #     json.dump(j, outfile)
            except:
                j = None
            if j:
                try:
                    li = j['props']['pageProps']['initialData']['data']['product']
                    # cat1 = li["category"]["path"][1]["name"]
                    # cat2 = li["category"]["path"][2]["name"]
                    upc = li["upc"]
                    today = datetime.today().date()
                    formatted_today = today.strftime('%Y-%m-%d')
                    data = {
                        "url": row[11],
                        "upc": trim_and_pad(upc),
                        "crawl_date": formatted_today}
                    print(data)
                    worksheet["M" + str(idx)].value = trim_and_pad(upc)
                    dataset.append(data)
                    missing = {
                            'p_id':int(row[11].split('?')[0].split('/')[-1]),
                            'upc': li["upc"],
                            'url': row[11]
                        }
                    missing_dataset.append(missing)
                except:
                    pass
    batch_insert(missing_dataset,insert_query)        
    df = pd.DataFrame(dataset)
    df.to_excel("Walmart_missing_60564.xlsx", index=False)
    workbook.save('Walmart__full_60564.xlsx')
# cat_links()
# p_links()
p_data()



